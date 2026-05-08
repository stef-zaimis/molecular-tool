from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from molecular_diagnosis.constants import COLORS
from molecular_diagnosis.core import (
    compute_match_score,
    compute_similarity,
    extract_sites,
)

from molecular_diagnosis.models import PunishmentResult


def autosize_columns(ws: Worksheet) -> None:
    for column in ws.columns:
        max_len = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_len + 2


def build_sheet(
    workbook: Workbook,
    name: str,
    sequences: dict[str, str],
    ref_id: str,
    sites: list[int] | tuple[int, ...],
    target_string: str,
) -> None:
    ws = workbook.create_sheet(name)

    header = ["ID"] + [site + 1 for site in sites] + ["matches", "% similarity", "avg similarity"]
    ws.append(header)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "B3"

    ref_states = extract_sites(sequences[ref_id], sites)

    rows = []

    for sequence_id, sequence in sequences.items():
        if sequence_id != ref_id and target_string in sequence_id:
            continue

        states = extract_sites(sequence, sites)
        match_score = compute_match_score(ref_states, states)
        similarity = compute_similarity(ref_states, states)

        rows.append((sequence_id, states, match_score, similarity))

    ref_row = next(row for row in rows if row[0] == ref_id)
    non_ref_rows = [row for row in rows if row[0] != ref_id]

    avg_similarity = sum(row[3] for row in non_ref_rows) / len(non_ref_rows)

    non_ref_rows = sorted(non_ref_rows, key=lambda row: (-row[2], -row[3], row[0]))
    final_rows = [ref_row] + non_ref_rows

    for row_index, (sequence_id, states, match_score, similarity) in enumerate(final_rows, start=2):
        avg_value = round(avg_similarity * 100, 2) if row_index == 2 else ""

        ws.append(
            [
                sequence_id,
                *states,
                round(match_score, 2),
                round(similarity * 100, 2),
                avg_value,
            ]
        )

        for column_index, base in enumerate(states, start=2):
            color = COLORS.get(base, "FFFFFF")

            ws.cell(row_index, column_index).fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid",
            )

    autosize_columns(ws)


def write_excel_report(
    output_path: str | Path,
    sequences: dict[str, str],
    ref_id: str,
    full_sites: list[int],
    target_string: str,
    best_gap_sites: tuple[int, ...] | None,
    best_avg_sites: tuple[int, ...] | None,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    build_sheet(
        workbook=workbook,
        name="Full",
        sequences=sequences,
        ref_id=ref_id,
        sites=full_sites,
        target_string=target_string,
    )

    if best_gap_sites is not None:
        build_sheet(
            workbook=workbook,
            name="Gap5",
            sequences=sequences,
            ref_id=ref_id,
            sites=best_gap_sites,
            target_string=target_string,
        )

    if best_avg_sites is not None:
        build_sheet(
            workbook=workbook,
            name="Avg5",
            sequences=sequences,
            ref_id=ref_id,
            sites=best_avg_sites,
            target_string=target_string,
        )

    workbook.save(output_path)

def write_punishment_excel_report(
    output_path: str | Path,
    focal_headers: list[str],
    alignment_length: int,
    punishment_result: PunishmentResult,
) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Punishments"

    header = [
        "ID",
        "PS",
        "PS/bp",
        "BD",
        "INS",
        "PRL",
        "POLY_PS",
        "BAL_PS",
        "INS_PS",
        "PRL_PS",
    ]

    ws.append(header)

    header_fill = PatternFill(
        start_color="595959",
        end_color="595959",
        fill_type="solid",
    )

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:J{len(focal_headers) + 1}"

    sorted_headers = sorted(
        focal_headers,
        key=lambda sequence_id: (
            -punishment_result.total_scores.get(sequence_id, 0.0),
            sequence_id,
        ),
    )

    for sequence_id in sorted_headers:
        total_score = punishment_result.total_scores.get(sequence_id, 0.0)

        if alignment_length > 0:
            score_per_bp = total_score / alignment_length
        else:
            score_per_bp = 0.0

        ws.append(
            [
                sequence_id,
                total_score,
                score_per_bp,
                punishment_result.bd_counts.get(sequence_id, 0),
                punishment_result.ins_counts.get(sequence_id, 0),
                punishment_result.prl_counts.get(sequence_id, 0),
                punishment_result.polymorphism_scores.get(sequence_id, 0.0),
                punishment_result.balancing_scores.get(sequence_id, 0.0),
                punishment_result.insertion_scores.get(sequence_id, 0.0),
                punishment_result.prolongation_scores.get(sequence_id, 0.0),
            ]
        )

    for row in ws.iter_rows(min_row=2):
        row[1].number_format = "0.000"
        row[2].number_format = "0.000000"
        row[3].number_format = "0"
        row[4].number_format = "0"
        row[5].number_format = "0"
        row[6].number_format = "0.000"
        row[7].number_format = "0.000"
        row[8].number_format = "0.000"
        row[9].number_format = "0.000"

        for cell in row[1:]:
            cell.alignment = Alignment(horizontal="center")

    autosize_columns(ws)

    workbook.save(output_path)