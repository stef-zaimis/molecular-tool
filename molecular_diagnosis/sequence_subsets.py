from dataclasses import dataclass
from pathlib import Path
from collections.abc import Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class SequenceSubsetGroup:
    representative_id: str
    representative_ungapped_sequence: str
    members: tuple[tuple[str, str], ...]

    @property
    def count(self) -> int:
        return len(self.members)


def ungap(sequence: str) -> str:
    return sequence.replace("-", "")


def group_sequence_subsets(
    sequences: Mapping[str, str],
) -> list[SequenceSubsetGroup]:
    """
    Group focal sequences by ungapped substring containment.

    Representative = longest ungapped sequence in the group.
    Shorter ungapped sequences that occur inside the representative are listed
    underneath as subsets.
    """
    entries = [
        (header, ungap(sequence))
        for header, sequence in sequences.items()
    ]

    entries.sort(key=lambda item: (-len(item[1]), item[0]))

    groups: list[dict[str, object]] = []

    for header, ungapped_sequence in entries:
        placed = False

        for group in groups:
            representative_sequence = group["rep_seq"]

            if not isinstance(representative_sequence, str):
                raise TypeError("Invalid representative sequence.")

            if ungapped_sequence and ungapped_sequence in representative_sequence:
                members = group["members"]

                if not isinstance(members, list):
                    raise TypeError("Invalid group members.")

                members.append((header, ungapped_sequence))
                placed = True
                break

            if not ungapped_sequence and not representative_sequence:
                members = group["members"]

                if not isinstance(members, list):
                    raise TypeError("Invalid group members.")

                members.append((header, ungapped_sequence))
                placed = True
                break

        if not placed:
            groups.append(
                {
                    "rep_id": header,
                    "rep_seq": ungapped_sequence,
                    "members": [(header, ungapped_sequence)],
                }
            )

    output: list[SequenceSubsetGroup] = []

    for group in groups:
        representative_id = group["rep_id"]
        representative_sequence = group["rep_seq"]
        members = group["members"]

        if not isinstance(representative_id, str):
            raise TypeError("Invalid representative ID.")
        if not isinstance(representative_sequence, str):
            raise TypeError("Invalid representative sequence.")
        if not isinstance(members, list):
            raise TypeError("Invalid group members.")

        sorted_members = sorted(
            members,
            key=lambda item: (-len(item[1]), item[0]),
        )

        output.append(
            SequenceSubsetGroup(
                representative_id=representative_id,
                representative_ungapped_sequence=representative_sequence,
                members=tuple(sorted_members),
            )
        )

    return output


def autosize_columns(ws: Worksheet, *, max_width: int = 100) -> None:
    for column in ws.columns:
        max_len = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_len + 2, max_width)


def excel_safe_sequence(sequence: str) -> str:
    """
    Excel cells have a 32,767-character limit.
    Barcode/COI-style sequences will normally be far below this.
    """
    max_len = 32767

    if len(sequence) <= max_len:
        return sequence

    suffix = f"... [TRUNCATED; full ungapped length={len(sequence)}]"
    return sequence[: max_len - len(suffix)] + suffix


def write_sequence_subset_excel_report(
    output_path: str | Path,
    sequence_subset_groups: list[SequenceSubsetGroup],
) -> None:
    output_path = Path(output_path)

    workbook = Workbook()
    ws = workbook.active
    ws.title = "UniqueSubsets"

    header = [
        "Group",
        "Row type",
        "Representative ID",
        "Group size",
        "Member ID",
        "Ungapped length",
        "Ungapped sequence",
    ]
    ws.append(header)

    header_fill = PatternFill(
        start_color="595959",
        end_color="595959",
        fill_type="solid",
    )

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    group_fill = PatternFill(
        start_color="D9EAD3",
        end_color="D9EAD3",
        fill_type="solid",
    )

    for group_index, group in enumerate(sequence_subset_groups, start=1):
        group_label = f"Group {group_index}"

        representative_row = ws.max_row + 1

        ws.append(
            [
                group_label,
                "representative",
                group.representative_id,
                group.count,
                group.representative_id,
                len(group.representative_ungapped_sequence),
                excel_safe_sequence(group.representative_ungapped_sequence),
            ]
        )

        for cell in ws[representative_row]:
            cell.font = Font(bold=True)
            cell.fill = group_fill

        for member_id, member_sequence in group.members:
            if member_id == group.representative_id:
                continue

            ws.append(
                [
                    group_label,
                    "subset",
                    group.representative_id,
                    "",
                    member_id,
                    len(member_sequence),
                    excel_safe_sequence(member_sequence),
                ]
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"

    autosize_columns(ws)

    workbook.save(output_path)