import tkinter as tk
from tkinter import filedialog, messagebox
import tkinter.font as tkfont

from itertools import combinations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

TXT_OUTPUT_BASENAME = "DMCs_output.txt"
XLSX_OUTPUT_BASENAME = "comparison_output.xlsx"

COLORS = {
    "A": "90EE90",
    "C": "87CEFA",
    "T": "FF7F7F",
    "G": "FFA500",
    "N": "C0C0C0",
    "R": "FFD27F",
    "Y": "D8BFD8",
    "W": "FFF2B2",
    "M": "BFEFFF",
    "-": "E0E0E0",
}

IUPAC = {
    "A": {"A"},
    "C": {"C"},
    "G": {"G"},
    "T": {"T"},
    "R": {"A", "G"},
    "Y": {"C", "T"},
    "W": {"A", "T"},
    "M": {"A", "C"},
    "N": {"A", "C", "G", "T"},
    "-": set(),
}


# ---------------- UTIL ----------------

def next_available_filename(path):
    path = Path(path)
    parent = path.parent
    stem = path.stem
    suffix = path.suffix

    existing = list(parent.glob(f"{stem}*{suffix}"))

    if not (parent / f"{stem}{suffix}").exists():
        return str(parent / f"{stem}{suffix}")

    numbers = []
    for f in existing:
        name = f.stem

        if name == stem:
            numbers.append(1)
        elif name.startswith(stem + "(") and name.endswith(")"):
            try:
                num = int(name[len(stem) + 1:-1])
                numbers.append(num)
            except ValueError:
                pass

    next_num = max(numbers) + 1 if numbers else 2
    return str(parent / f"{stem}({next_num}){suffix}")


def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2


def parse_fasta(path):
    sequences = {}

    with open(path) as f:
        header = None
        seq_chunks = []

        for line in f:
            line = line.strip()

            if not line:
                continue

            if line.startswith(">"):
                if header:
                    sequences[header] = "".join(seq_chunks).upper()
                header = line[1:]
                seq_chunks = []
            else:
                seq_chunks.append(line)

        if header:
            sequences[header] = "".join(seq_chunks).upper()

    return sequences


def extract_sites(seq, sites):
    return [seq[i] for i in sites]


def score_state(ref, query):
    possibilities = IUPAC.get(query, set())

    if not possibilities:
        return 0.0

    if ref in possibilities:
        return 1.0 / len(possibilities)

    return 0.0


def compute_similarity(ref_states, query_states):
    return sum(score_state(r, q) for r, q in zip(ref_states, query_states)) / len(ref_states)


def compute_match_score(ref_states, query_states):
    return sum(score_state(r, q) for r, q in zip(ref_states, query_states))


# ---------------- CORE ----------------

def compute_metrics(sequences, ref_id, sites, target_string):
    ref_states = extract_sites(sequences[ref_id], sites)

    sims = []
    max_sim = -1.0

    for h, seq in sequences.items():
        if h == ref_id or target_string in h:
            continue

        q_states = extract_sites(seq, sites)
        sim = compute_similarity(ref_states, q_states)

        sims.append(sim)

        if sim > max_sim:
            max_sim = sim

    avg_sim = sum(sims) / len(sims)
    return max_sim, avg_sim, 1 - max_sim


def find_dmc_information(sequences, target_string):
    all_seqs = list(sequences.values())
    focal = [s for h, s in sequences.items() if target_string in h]
    non_focal = [s for h, s in sequences.items() if target_string not in h]

    length = len(focal[0])

    fixed = {}
    skipped_non_acgt = 0

    for i in range(length):
        col = [s[i] for s in focal]

        if any(c not in "ACGT" for c in col):
            skipped_non_acgt += 1
            continue

        if len(set(col)) == 1:
            fixed[i] = col[0]

    candidates = {}
    globally_conserved_removed = 0

    for i, b in fixed.items():
        col = [s[i] for s in all_seqs if s[i] in "ACGT"]

        if len(set(col)) > 1:
            candidates[i] = b
        else:
            globally_conserved_removed += 1

    sites = list(candidates.keys())

    singles = []
    for i in sites:
        if all(s[i] != candidates[i] for s in non_focal):
            singles.append(i)

    total_pairs_tested = 0
    pairs = []

    for i in range(len(sites)):
        for j in range(i + 1, len(sites)):
            total_pairs_tested += 1

            a, b = sites[i], sites[j]

            if all(not (s[a] == candidates[a] and s[b] == candidates[b]) for s in non_focal):
                pairs.append((a, b))

    unique = sorted(set(x for p in pairs for x in p))

    return {
        "single": singles,
        "pairs": pairs,
        "unique": unique,
        "fixed_count": len(fixed),
        "skipped_non_acgt": skipped_non_acgt,
        "globally_conserved_removed": globally_conserved_removed,
        "candidate_count": len(candidates),
        "pairs_tested": total_pairs_tested,
    }


def format_diag(sites, seq):
    if not sites:
        return "None"

    return ", ".join(f"{i + 1}:{seq[i]}" for i in sites)


def build_sheet(wb, name, sequences, ref_id, sites, target_string):
    ws = wb.create_sheet(name)

    header = ["ID"] + [i + 1 for i in sites] + ["matches", "% similarity", "avg similarity"]
    ws.append(header)

    for c in ws[1]:
        c.font = Font(bold=True)

    ws.freeze_panes = "B3"

    ref_states = extract_sites(sequences[ref_id], sites)

    rows = []

    for h, seq in sequences.items():
        if h != ref_id and target_string in h:
            continue

        states = extract_sites(seq, sites)
        m = compute_match_score(ref_states, states)
        sim = compute_similarity(ref_states, states)

        rows.append((h, states, m, sim))

    ref_row = next(r for r in rows if r[0] == ref_id)
    non_ref = [r for r in rows if r[0] != ref_id]

    avg = sum(r[3] for r in non_ref) / len(non_ref)

    non_ref = sorted(non_ref, key=lambda x: (-x[2], -x[3], x[0]))
    final = [ref_row] + non_ref

    for i, (h, states, m, sim) in enumerate(final, start=2):
        avg_val = round(avg * 100, 2) if i == 2 else ""
        ws.append([h, *states, round(m, 2), round(sim * 100, 2), avg_val])

        for j, b in enumerate(states, start=2):
            ws.cell(i, j).fill = PatternFill(
                start_color=COLORS.get(b, "FFFFFF"),
                end_color=COLORS.get(b, "FFFFFF"),
                fill_type="solid"
            )

    autosize_columns(ws)


# ---------------- PIPELINE ----------------

def run_pipeline_core(fasta_path, target_string, output_dir):
    fasta_path = str(fasta_path).strip()
    target_string = str(target_string).strip()
    output_dir = str(output_dir).strip()

    if not fasta_path:
        raise ValueError("No FASTA file selected.")

    if not target_string:
        raise ValueError("No identifier string entered.")

    if not output_dir:
        raise ValueError("No output directory selected.")

    fasta_path_obj = Path(fasta_path)
    output_dir_obj = Path(output_dir)

    if not fasta_path_obj.exists():
        raise ValueError("FASTA file does not exist.")

    if not output_dir_obj.exists():
        raise ValueError("Output directory does not exist.")

    sequences = parse_fasta(fasta_path)

    if not sequences:
        raise ValueError("No sequences were read from the FASTA file.")

    lengths = {len(seq) for seq in sequences.values()}
    if len(lengths) != 1:
        raise ValueError("Sequences are not all the same length. Input must be an aligned FASTA.")

    focal_headers = [h for h in sequences if target_string in h]
    non_focal_headers = [h for h in sequences if target_string not in h]

    if len(focal_headers) == 0:
        raise ValueError("No sequences matched the identifier.")

    if len(non_focal_headers) == 0:
        raise ValueError("All sequences match the identifier; need contrast set.")

    ref_id = focal_headers[0]
    ref_seq = sequences[ref_id]

    dmc = find_dmc_information(sequences, target_string)
    sites = dmc["unique"]

    total_5site_combos = 0
    best_gap = None
    best_gap_set = None
    best_avg = None
    best_avg_set = None

    if len(sites) >= 5:
        for combo in combinations(sites, 5):
            total_5site_combos += 1
            m, a, g = compute_metrics(sequences, ref_id, combo, target_string)

            if best_gap is None or m < best_gap:
                best_gap, best_gap_set = m, combo

            if best_avg is None or a < best_avg:
                best_avg, best_avg_set = a, combo

    txt_output_path = next_available_filename(output_dir_obj / TXT_OUTPUT_BASENAME)
    xlsx_output_path = next_available_filename(output_dir_obj / XLSX_OUTPUT_BASENAME)

    with open(txt_output_path, "w") as f:
        f.write("Molecular diagnostic report\n\n")

        f.write("Input settings:\n")
        f.write(f"FASTA file: {fasta_path}\n")
        f.write(f"Output directory: {output_dir}\n")
        f.write(f"Focal identifier string: {target_string}\n\n")

        f.write("Sequence summary:\n")
        f.write(f"Total sequences read: {len(sequences)}\n")
        f.write(f"Alignment length: {next(iter(lengths))}\n")
        f.write(f"Focal sequences: {len(focal_headers)}\n")
        f.write(f"Non-focal sequences: {len(non_focal_headers)}\n")
        f.write(f"Reference sequence selected: {ref_id}\n\n")

        f.write("Run diagnostics:\n")
        f.write(f"Sites skipped because focal column contains non-ACGT: {dmc['skipped_non_acgt']}\n")
        f.write(f"Sites fixed in focal set (strict A/C/G/T only): {dmc['fixed_count']}\n")
        f.write(f"Fixed sites removed as globally conserved: {dmc['globally_conserved_removed']}\n")
        f.write(f"Candidate diagnostic sites retained: {dmc['candidate_count']}\n")
        f.write(f"1-site diagnoses found: {len(dmc['single'])}\n")
        f.write(f"2-site combinations tested: {dmc['pairs_tested']}\n")
        f.write(f"Valid 2-site diagnostic combinations recovered: {len(dmc['pairs'])}\n")
        f.write(f"Unique sites participating in valid 2-site combinations: {len(sites)}\n")
        f.write(f"5-site combinations tested: {total_5site_combos}\n\n")

        f.write("1-site diagnosis:\n")
        if not dmc["single"]:
            f.write("None found\n\n")
        else:
            f.write(format_diag(dmc["single"], ref_seq) + "\n\n")

        f.write("2-site diagnostic combinations:\n")
        if not dmc["pairs"]:
            f.write("None found\n")
        else:
            for a, b in dmc["pairs"]:
                f.write(f"[{a + 1}:{ref_seq[a]}, {b + 1}:{ref_seq[b]}]\n")

        f.write(f"\nTotal: {len(dmc['pairs'])}\n\n")

        f.write("Unique diagnostic sites:\n")
        f.write(format_diag(sites, ref_seq) + "\n\n")

        f.write(f"Full diagnosis ({len(sites)} sites):\n")
        f.write(format_diag(sites, ref_seq) + "\n\n")

        f.write("5-site diagnosis (for similarity gap optimisation):\n")
        if best_gap_set is None:
            f.write("Not computed: fewer than 5 unique diagnostic sites.\n\n")
        else:
            f.write(format_diag(best_gap_set, ref_seq) + "\n\n")

        f.write("5-site diagnosis (for average similarity optimisation):\n")
        if best_avg_set is None:
            f.write("Not computed: fewer than 5 unique diagnostic sites.\n\n")
        else:
            f.write(format_diag(best_avg_set, ref_seq) + "\n\n")

    wb = Workbook()
    wb.remove(wb.active)

    build_sheet(wb, "Full", sequences, ref_id, sites, target_string)

    if best_gap_set is not None:
        build_sheet(wb, "Gap5", sequences, ref_id, best_gap_set, target_string)

    if best_avg_set is not None:
        build_sheet(wb, "Avg5", sequences, ref_id, best_avg_set, target_string)

    wb.save(xlsx_output_path)

    return txt_output_path, xlsx_output_path


# ---------------- GUI ----------------

def launch_gui():
    root = tk.Tk()
    root.title("Molecular Diagnosis Tool")
    root.geometry("650x360")

    fasta_var = tk.StringVar()
    target_var = tk.StringVar()
    output_dir_var = tk.StringVar()

    def view_fasta():
        fasta_path = fasta_var.get().strip()

        if not fasta_path:
            messagebox.showerror("Error", "Please select a FASTA file first.")
            return

        try:
            sequences = parse_fasta(fasta_path)
        except Exception as e:
            messagebox.showerror("Error", f"Unable to read the file.\n\nError type: {type(e).__name__}\nDetails: {e}")
            return

        if not sequences:
            messagebox.showerror("Error", "Unable to read the file.\n\nError type: EmptyFASTA\nDetails: No sequences were read from the FASTA file.")
            return

        viewer = tk.Toplevel(root)
        viewer.title("FASTA Viewer")
        viewer.geometry("1200x500")

        font = tkfont.nametofont("TkDefaultFont")
        bold_font = font.copy()
        bold_font.configure(weight="bold")

        name_width = max(
            font.measure("Sequence name"),
            max(font.measure(header) for header in sequences.keys())
        ) + 6

        base_width = font.measure("W") + 6
        row_height = 24
        seq_length = max(len(sequence) for sequence in sequences.values())

        position_digits = len(str(seq_length))
        digit_spacing = font.metrics("linespace") * 0.75
        header_height = int(position_digits * digit_spacing + 8)

        outer = tk.Frame(viewer)
        outer.pack(fill="both", expand=True)

        corner_canvas = tk.Canvas(outer, width=name_width, height=header_height, highlightthickness=0)
        top_canvas = tk.Canvas(outer, height=header_height, highlightthickness=0)
        name_canvas = tk.Canvas(outer, width=name_width, highlightthickness=0)
        seq_canvas = tk.Canvas(outer, highlightthickness=0)

        y_scroll = tk.Scrollbar(outer, orient="vertical")
        x_scroll = tk.Scrollbar(outer, orient="horizontal")

        corner_canvas.grid(row=0, column=0, sticky="nsew")
        top_canvas.grid(row=0, column=1, sticky="ew")
        name_canvas.grid(row=1, column=0, sticky="ns")
        seq_canvas.grid(row=1, column=1, sticky="nsew")
        y_scroll.grid(row=1, column=2, sticky="ns")
        x_scroll.grid(row=2, column=1, sticky="ew")

        outer.grid_rowconfigure(1, weight=1)
        outer.grid_columnconfigure(1, weight=1)

        def yview(*args):
            name_canvas.yview(*args)
            seq_canvas.yview(*args)

        def xview(*args):
            top_canvas.xview(*args)
            seq_canvas.xview(*args)

        y_scroll.config(command=yview)
        x_scroll.config(command=xview)

        name_canvas.config(yscrollcommand=y_scroll.set)
        seq_canvas.config(xscrollcommand=x_scroll.set, yscrollcommand=y_scroll.set)

        corner_canvas.create_rectangle(0, 0, name_width, header_height, fill="white", outline="gray")
        corner_canvas.create_text(3, header_height / 2, text="Sequence name", anchor="w", font=bold_font)

        for pos in range(seq_length):
            x1 = pos * base_width
            x2 = x1 + base_width
            label = f"{pos + 1:0{position_digits}d}"

            top_canvas.create_rectangle(x1, 0, x2, header_height, fill="white", outline="gray")

            center_x = (x1 + x2) / 2
            start_y = (header_height - ((position_digits - 1) * digit_spacing)) / 2

            seen_nonzero = False

            for digit_index, digit in enumerate(label):
                y = start_y + digit_index * digit_spacing

                if digit != "0":
                    seen_nonzero = True

                digit_fill = "light gray" if digit == "0" and not seen_nonzero else "black"

                top_canvas.create_text(
                    center_x,
                    y,
                    text=digit,
                    anchor="center",
                    font=font,
                    fill=digit_fill
                )

        row_text_items = {}
        hovered_row = None

        for row_index, (header, sequence) in enumerate(sequences.items()):
            y1 = row_index * row_height
            y2 = y1 + row_height

            name_canvas.create_rectangle(0, y1, name_width, y2, fill="white", outline="gray")
            name_text = name_canvas.create_text(3, (y1 + y2) / 2, text=header, anchor="w", font=font)

            padded_sequence = sequence.ljust(seq_length, "-")

            row_text_items[row_index] = [(name_canvas, name_text)]

            for pos, base in enumerate(padded_sequence):
                x1 = pos * base_width
                x2 = x1 + base_width
                fill = "#" + COLORS.get(base, "FFFFFF")

                seq_canvas.create_rectangle(x1, y1, x2, y2, fill=fill, outline="gray")
                base_text = seq_canvas.create_text(
                    (x1 + x2) / 2,
                    (y1 + y2) / 2,
                    text=base,
                    anchor="center",
                    font=font
                )

                row_text_items[row_index].append((seq_canvas, base_text))

        total_height = len(sequences) * row_height
        total_width = seq_length * base_width

        top_canvas.config(scrollregion=(0, 0, total_width, header_height))
        name_canvas.config(scrollregion=(0, 0, name_width, total_height))
        seq_canvas.config(scrollregion=(0, 0, total_width, total_height))

        def set_row_font(row_index, selected):
            if row_index not in row_text_items:
                return

            chosen_font = bold_font if selected else font

            for canvas, item in row_text_items[row_index]:
                canvas.itemconfigure(item, font=chosen_font)

        def clear_hover():
            nonlocal hovered_row

            if hovered_row is not None:
                set_row_font(hovered_row, False)
                hovered_row = None

        def row_from_event(canvas, event):
            y = canvas.canvasy(event.y)
            row_index = int(y // row_height)

            if row_index < 0 or row_index >= len(sequences):
                return None

            return row_index

        def on_row_motion(canvas, event):
            nonlocal hovered_row

            row_index = row_from_event(canvas, event)

            if row_index is None:
                clear_hover()
                return "break"

            if row_index != hovered_row:
                clear_hover()
                set_row_font(row_index, True)
                hovered_row = row_index

            return "break"

        def scroll_vertical(units):
            name_canvas.yview_scroll(units, "units")
            seq_canvas.yview_scroll(units, "units")
            clear_hover()

        def scroll_horizontal(units):
            top_canvas.xview_scroll(units, "units")
            seq_canvas.xview_scroll(units, "units")
            clear_hover()

        def on_mousewheel(event):
            if event.state & 0x0004:
                scroll_horizontal(int(-1 * (event.delta / 120)))
            else:
                scroll_vertical(int(-1 * (event.delta / 120)))
            return "break"

        def on_linux_scroll_up(event):
            if event.state & 0x0004:
                scroll_horizontal(-3)
            else:
                scroll_vertical(-3)
            return "break"

        def on_linux_scroll_down(event):
            if event.state & 0x0004:
                scroll_horizontal(3)
            else:
                scroll_vertical(3)
            return "break"

        name_canvas.bind("<Motion>", lambda event: on_row_motion(name_canvas, event))
        seq_canvas.bind("<Motion>", lambda event: on_row_motion(seq_canvas, event))

        for canvas in (corner_canvas, top_canvas, name_canvas, seq_canvas):
            canvas.bind("<MouseWheel>", on_mousewheel)
            canvas.bind("<Button-4>", on_linux_scroll_up)
            canvas.bind("<Button-5>", on_linux_scroll_down)
            canvas.bind("<Leave>", lambda event: clear_hover())

        viewer.bind("<Leave>", lambda event: clear_hover())

    def browse_fasta():
        path = filedialog.askopenfilename(
            initialdir=str(Path.home()),
            title="Select aligned FASTA file",
            filetypes=[
                ("FASTA files", "*.fasta *.fa *.fna"),
                ("All files", "*.*")
            ]
        )

        if path:
            fasta_var.set(path)

    def browse_output():
        folder = filedialog.askdirectory(
            initialdir=str(Path.home()),
            title="Select output folder"
        )

        if folder:
            output_dir_var.set(folder)

    def use_fasta_location():
        fasta_path = fasta_var.get().strip()

        if not fasta_path:
            messagebox.showerror("Error", "Select a FASTA file first.")
            return

        output_dir_var.set(str(Path(fasta_path).parent))

    def run():
        fasta_path = fasta_var.get().strip()
        target_string = target_var.get().strip()
        output_dir = output_dir_var.get().strip()

        if not fasta_path:
            messagebox.showerror("Error", "Please select a FASTA file.")
            return

        if not target_string:
            messagebox.showerror("Error", "Please enter an identifier string.")
            return

        if not output_dir:
            messagebox.showerror("Error", "Please select an output directory.")
            return

        try:
            txt_path, xlsx_path = run_pipeline_core(fasta_path, target_string, output_dir)
            messagebox.showinfo(
                "Done",
                f"Analysis completed.\n\nText output:\n{txt_path}\n\nExcel output:\n{xlsx_path}"
            )
        except Exception as e:
            messagebox.showerror("Error", str(e))

    tk.Label(root, text="Aligned FASTA file").pack(pady=(10, 0))
    tk.Entry(root, textvariable=fasta_var, width=80).pack(padx=10)

    fasta_button_frame = tk.Frame(root)
    fasta_button_frame.pack(pady=4)

    tk.Button(fasta_button_frame, text="Browse FASTA", command=browse_fasta)\
        .grid(row=0, column=0, padx=5)

    tk.Button(fasta_button_frame, text="View FASTA", command=view_fasta)\
        .grid(row=0, column=1, padx=5)

    tk.Label(root, text="Focal identifier string").pack(pady=(10, 0))
    tk.Entry(root, textvariable=target_var, width=80).pack(padx=10)

    tk.Label(root, text="Output directory").pack(pady=(10, 0))
    tk.Entry(root, textvariable=output_dir_var, width=80).pack(padx=10)

    # container for horizontal buttons with "or"
    btn_frame = tk.Frame(root)
    btn_frame.pack(pady=4)

    tk.Button(btn_frame, text="Browse Output Folder", command=browse_output)\
        .grid(row=0, column=0, padx=5)

    tk.Label(btn_frame, text="or")\
        .grid(row=0, column=1, padx=5)

    tk.Button(btn_frame, text="Use FASTA Location", command=use_fasta_location)\
        .grid(row=0, column=2, padx=5)
    tk.Button(root, text="Run", command=run).pack(pady=15)

    root.mainloop()


if __name__ == "__main__":
    launch_gui()
